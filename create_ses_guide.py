"""
SES Model - ses_main.dat rehber xlsx oluşturucu
Tüm input türlerini, açıklamalarını ve yeni ülke/kaynak ekleme yönlendirmelerini içerir.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Renk paleti ──────────────────────────────────────────────────────────────
C_HEADER_DARK  = "1F4E79"   # koyu lacivert  (başlık satırları)
C_HEADER_MED   = "2E75B6"   # orta mavi      (alt başlıklar)
C_HEADER_LIGHT = "BDD7EE"   # açık mavi      (grup rengi)
C_YELLOW       = "FFD966"   # sarı            (uyarı / önemli)
C_GREEN        = "E2EFDA"   # açık yeşil      (örnek veriler)
C_ORANGE       = "FCE4D6"   # turuncu         (değiştirilecek alanlar)
C_GREY         = "F2F2F2"   # açık gri        (açıklama satırları)
C_WHITE        = "FFFFFF"
C_RED_LIGHT    = "FFCCCC"   # açık kırmızı    (uyarı)

def make_font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(thick=False):
    s = "medium" if thick else "thin"
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)

def header_cell(ws, row, col, value, bg=C_HEADER_DARK, fg="FFFFFF",
                bold=True, size=10, wrap=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = make_font(bold=bold, color=fg, size=size)
    c.fill      = make_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    c.border    = make_border()
    return c

def data_cell(ws, row, col, value, bg=C_WHITE, bold=False,
              align="left", wrap=True, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = make_font(bold=bold, color=color)
    c.fill      = make_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border    = make_border()
    return c

def section_title(ws, row, col, text, span, bg=C_HEADER_MED):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = make_font(bold=True, color="FFFFFF", size=11)
    c.fill      = make_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            wrap_text=False)
    c.border    = make_border(thick=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                ln = max(len(str(line)) for line in str(cell.value).split("\n"))
                best = min(max(best, ln + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 – README
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
c = ws.cell(row=1, column=1,
            value="Swiss EnergyScope (SES) – ses_main.dat Rehberi")
c.font      = make_font(bold=True, color="FFFFFF", size=16)
c.fill      = make_fill(C_HEADER_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

readme_rows = [
    ("AMAÇ", "Bu dosya ses_main.dat içindeki tüm input türlerini Türkçe açıklamalarıyla belgeler ve yeni bir ülke modellemek ya da yeni enerji kaynağı eklemek için adım adım yönlendirme sağlar."),
    ("BİRİMLER", "Enerji: GWh | Güç: GW | Maliyet: MCHF (Milyon İsviçre Frangı) | Süre: h | Yolcu taşımacılığı: Mpkm | Yük taşımacılığı: Mtkm"),
    ("DÖNEMLER", "Model 12 dönemle çalışır (aylık çözünürlük). Her dönem tipik olarak 1 ay = ~720–744 saate karşılık gelir."),
    ("TEMEL KAVRAMLAR", "RESOURCES: Satın alınabilir birincil enerji kaynakları\nTECHNOLOGIES: Kaynakları nihai kullanıma dönüştüren teknolojiler\nEND_USES: Karşılanması gereken nihai enerji talep kategorileri\nLAYERS: Enerji taşıyıcıları (elektrik, doğal gaz, ısı vb.)\nSTORAGE: Depolama teknolojileri"),
    ("SAYFALAR", "SETS              → Model setleri ve açıklamaları\nPARAMETRELER      → Tüm parametreler özet tablosu\nTALEP_VERISI      → end_uses_demand_year\nLAYERS_IN_OUT     → Teknoloji girdi/çıktı katmanları\nTEKNOLOJI_PARAMS  → Maliyet, ömür, kapasite faktörleri\nKAYNAK_VERISI     → Kaynak kullanılabilirlik, maliyet, emisyon\nZAMAN_SERISİ      → Kapasite faktörleri & aylık parametreler\nDEPOLAMA          → Depolama teknoloji parametreleri\nYENI_ULKE_REHBERI → Yeni ülke adım adım kılavuzu\nKOMUR_EKLEME      → Kömür kaynağı/teknoloji ekleme kılavuzu\nAYLIK_COZUNURLUK  → Aylık çözünürlük ayarları"),
]

ws.row_dimensions[2].height = 6  # boşluk

for i, (title, desc) in enumerate(readme_rows, start=3):
    ws.row_dimensions[i].height = 55
    c1 = ws.cell(row=i, column=1, value=title)
    c1.font      = make_font(bold=True, color=C_HEADER_DARK, size=10)
    c1.fill      = make_fill(C_HEADER_LIGHT)
    c1.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    c1.border    = make_border()
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    c2 = ws.cell(row=i, column=2, value=desc)
    c2.font      = make_font(size=10)
    c2.fill      = make_fill(C_WHITE)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border    = make_border()

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 90
for col in "CDEF":
    ws.column_dimensions[col].width = 5

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 – SETS
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("SETS")
ws2.sheet_view.showGridLines = False

section_title(ws2, 1, 1, "MODEL SETLERİ – AÇIKLAMALAR", 5)

headers = ["Set Adı", "Tip", "Mevcut Elemanlar (Özet)", "Türkçe Açıklama", "Yeni Ülke İçin Not"]
for j, h in enumerate(headers, 1):
    header_cell(ws2, 2, j, h, bg=C_HEADER_DARK)

sets_data = [
    ("PERIODS", "Zaman indeksi", "1 2 3 4 5 6 7 8 9 10 11 12",
     "12 aylık periyot. Her eleman 1 aya karşılık gelir. Ocak=1, Aralık=12.",
     "Değiştirme: Yıllık çözünürlük için '1' yap, saatlik için 8760 eleman ekle."),

    ("SECTORS", "Talep sektörü", "HOUSEHOLDS SERVICES INDUSTRY TRANSPORTATION",
     "Talep sektörleri: Konutlar, Hizmetler, Sanayi, Ulaştırma.",
     "Yeni ülkede aynı sektörler kullanılabilir; gerekirse AGRICULTURE eklenebilir."),

    ("END_USES_INPUT", "Nihai kullanım girişi",
     "ELECTRICITY LIGHTING HEAT_HIGH_T HEAT_LOW_T_SH HEAT_LOW_T_HW MOBILITY_PASSENGER MOBILITY_FREIGHT",
     "end_uses_demand_year tablosunun satır etiketleri. Her talep kategorisini tanımlar.",
     "Aynı kategoriler kullanılmalı; sadece sayısal değerler ülkeye göre değişir."),

    ("END_USES_CATEGORIES", "Nihai kullanım kategorisi",
     "ELECTRICITY HEAT_HIGH_T HEAT_LOW_T MOBILITY_PASSENGER MOBILITY_FREIGHT",
     "Üst düzey talep kategorileri. END_USES_INPUT ile birebir eşleşir.",
     "Değiştirilmez; modelin matematiksel yapısı bu kategorilere dayanır."),

    ("RESOURCES", "Birincil kaynak",
     "ELECTRICITY GASOLINE DIESEL BIOETHANOL BIODIESEL LFO LNG NG NG_CCS SNG WOOD COAL COAL_CCS URANIUM WASTE H2 ELEC_EXPORT",
     "Sisteme dışarıdan alınabilen enerji kaynakları. c_op (işletme maliyeti) ve avail (kullanılabilirlik) parametreleri bu set için tanımlanır.",
     "KÖMÜR ZATen dahil! Kömür kullanmak için avail>0 ve c_op>0 yapılmalı."),

    ("BIOFUELS_IMPORT", "Biyoyakıt alt seti", "BIOETHANOL BIODIESEL SNG",
     "İthal edilen biyoyakıtlar. Modelde bunların fiyatı fossil fuel fiyatına eşitlenir.",
     "Yeni ülkede yerli biyoyakıt varsa bu setten çıkar, ayrı parametre ver."),

    ("EXPORT", "İhracat alt seti", "ELEC_EXPORT",
     "Elektrik ihracatı için sanal kaynak. Negatif layers_in_out değeri ile elektriği 'tüketir'.",
     "İhracat için f_max parametresi ile üst sınır belirle."),

    ("END_USES_TYPES_OF_CATEGORY[...]", "Teknoloji–kategori eşleşmesi",
     "ELECTRICITY→{ELECTRICITY}, HEAT_HIGH_T→{HEAT_HIGH_T}, ...",
     "Her üst kategori hangi alt tipleri kapsar. Teknolojiler bu alt tiplere atanır.",
     "Yeni tip eklenirse hem bu set hem TECHNOLOGIES_OF_END_USES_TYPE güncellenmeli."),

    ("TECHNOLOGIES_OF_END_USES_TYPE[...]", "Teknoloji–end_use eşleşmesi",
     "ELECTRICITY→{NUCLEAR CCGT COAL_US PV WIND ...}, HEAT_HIGH_T→{IND_COGEN_GAS ...}",
     "Her nihai kullanım tipini hangi teknolojilerin karşılayabileceğini belirtir.",
     "Yeni teknoloji (örn. COAL_BOILER_NEW) eklerken ilgili alt tipe eklenmeli."),

    ("STORAGE_TECH", "Depolama teknolojisi", "PUMPED_HYDRO POWER2GAS",
     "Depolama yapabilen teknolojiler. storage_eff_in/out ve loss_coeff parametreleri buraya bağlı.",
     "Batarya depolama için BATTERY eklenebilir."),

    ("INFRASTRUCTURE", "Altyapı teknolojisi",
     "EFFICIENCY DHN GRID POWER2GAS_1 POWER2GAS_2 POWER2GAS_3 H2_ELECTROLYSIS H2_NG H2_BIOMASS GASIFICATION_SNG PYROLYSIS",
     "Doğrudan enerji dönüşümü yapmayan ama sistemi destekleyen altyapılar.",
     "DHN (district heating network) ve GRID (elektrik şebekesi) ülke bazında boyutlandırılmalı."),

    ("COGEN / BOILERS", "Sonuç yazdırma seti",
     "COGEN: kojenerasyon teknolojileri; BOILERS: kazan teknolojileri",
     "Sadece sonuç raporlama için kullanılır, optimizasyonu etkilemez.",
     "Yeni teknolojiler eklenirse raporlama için bu setlere de ekle."),
]

for i, row in enumerate(sets_data, start=3):
    bg = C_GREY if i % 2 == 0 else C_WHITE
    for j, val in enumerate(row, 1):
        data_cell(ws2, i, j, val, bg=bg)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 50
ws2.column_dimensions["D"].width = 55
ws2.column_dimensions["E"].width = 55
for r in range(3, 3+len(sets_data)):
    ws2.row_dimensions[r].height = 48

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 – PARAMETRELER (özet)
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("PARAMETRELER")
ws3.sheet_view.showGridLines = False
section_title(ws3, 1, 1, "TÜM PARAMETRELER – ÖZET TABLOSU", 6)

h3 = ["Parametre Adı", "Tanım", "Birim", "Set(ler)", "ses_main.dat'taki Değer(ler)", "Yeni Ülke Notu"]
for j, h in enumerate(h3, 1):
    header_cell(ws3, 2, j, h, bg=C_HEADER_DARK)

params_data = [
    # Genel
    ("i_rate", "Yıllık faiz/iskonto oranı",
     "– (oran)", "–",
     "0.03215 (≈ %3.2)",
     "Ülkenin risk-free faiz oranına göre ayarla. Türkiye için ~0.08–0.15 önerilir."),

    # Talep
    ("end_uses_demand_year", "Yıllık sektörel enerji talebi",
     "GWh/yıl", "END_USES_INPUT × SECTORS",
     "Bkz. TALEP_VERISI sayfası",
     "En kritik parametre. IEA/TEIAŞ verilerinden derlenmeli."),

    ("share_mobility_public_min/max", "Toplu taşımanın ulaştırmadaki min/max payı",
     "– (oran)", "–",
     "min=0.3, max=0.5",
     "Türkiye'de toplu taşıma payı düşük → min=0.15, max=0.4 öneri."),

    ("share_freight_train_min/max", "Demiryolu yük taşımacılığı payı",
     "– (oran)", "–",
     "min=0.4, max=0.6",
     "Türkiye'de karayolu baskın → min=0.05, max=0.3 öneri."),

    ("share_heat_dhn_min/max", "Bölgesel ısıtma ağı ısı payı",
     "– (oran)", "–",
     "min=0.1, max=0.3",
     "DHN altyapısı az → min=0.0, max=0.15 öneri."),

    # Zaman
    ("period_duration", "Her dönemin süresi",
     "saat", "PERIODS",
     "744 672 744 720 744 720 744 744 720 744 720 744",
     "Aylık çözünürlükte yılın aylarına göre saat sayısı (Ocak:744, Şubat:672 vb.)"),

    ("lighting_month", "Aydınlatma talebinin aylık dağılımı",
     "– (oran, Σ=1)", "PERIODS",
     "0.1240 0.1020 ... (bkz. ZAMAN_SERISİ)",
     "Güneş ışığı saatlerine göre ayarla; yaz aylarında düşük olmalı."),

    ("heating_month", "Isıtma talebinin aylık dağılımı",
     "– (oran, Σ=1)", "PERIODS",
     "0.1984 0.1654 ... (bkz. ZAMAN_SERISİ)",
     "Isıtma derece-günlerine (Heating Degree Days) göre ayarla."),

    # Katmanlar
    ("layers_in_out", "Teknoloji girdi(−)/çıktı(+) katsayıları (verimlilik)",
     "–", "TECHNOLOGIES × LAYERS",
     "Bkz. LAYERS_IN_OUT sayfası",
     "Teknoloji verimliliğini kodlar. Değer = çıktı/girdi. Girdi için negatif."),

    # Kaynaklar
    ("avail", "Yıllık maksimum kaynak kullanılabilirliği",
     "GWh/yıl", "RESOURCES",
     "WOOD=12279, WASTE=11142, diğerleri=1000000 (sınırsız)",
     "Yerli kaynaklar için gerçekçi üst sınır koy (ör. WOOD=Türkiye orman potansiyeli)."),

    ("gwp_op", "İşletme sırasında GHG emisyonu (karbondioksit eşdeğeri)",
     "tCO2eq/MWh", "RESOURCES",
     "NG=0.2666, COAL=0.4014, URANIUM=0.0039 ...",
     "IPCC emisyon faktörlerinden alınabilir. Ülkeye özgü grid emisyon faktörü önemli."),

    ("c_op", "Kaynak işletme maliyeti (aylık bazda)",
     "MCHF/GWh", "RESOURCES × PERIODS",
     "NG=0.0348, COAL=0.0302, URANIUM=0.0041 ...",
     "Enerji ithal fiyatlarını (USD/MWh) MCHF/GWh'e çevir: ×0.001×kur."),

    # Teknoloji parametreleri
    ("ref_size", "Referans tesis büyüklüğü (kapasite birimi için normalleştirme)",
     "GW", "TECHNOLOGIES",
     "NUCLEAR=1, PV=0.000003, CAR_GASOLINE=0.001 ...",
     "Modelin iç ölçeklendirmesi için; genellikle değiştirilmez."),

    ("c_inv", "Spesifik yatırım maliyeti",
     "MCHF/GW_ref", "TECHNOLOGIES",
     "NUCLEAR=5174.76, PV=1000, CCGT=824.41 ...",
     "IRENA/IEA maliyet veritabanından güncellenebilir. 2030 projeksiyon değerleri kullanılabilir."),

    ("c_maint", "Yıllık sabit bakım–işletme maliyeti",
     "MCHF/GW/yıl", "TECHNOLOGIES",
     "NUCLEAR=109.92, PV=15.88, CCGT=21.07 ...",
     "Genellikle c_inv'in %1–3'ü. Yerel işçilik maliyetiyle ölçeklendirilebilir."),

    ("gwp_constr", "İnşaat/üretim sırasındaki GHG emisyonu (yaşam döngüsü)",
     "tCO2eq/GW", "TECHNOLOGIES",
     "NUCLEAR=707.88, PV=2081.43, CCGT=183.79 ...",
     "LCA (yaşam döngüsü) analizinden alınır. İkincil önem taşır."),

    ("lifetime", "Teknoloji ekonomik ömrü",
     "yıl", "TECHNOLOGIES",
     "NUCLEAR=60, CCGT=25, PV=25, CAR=1 ...",
     "Araçlar için 1 yıl (zaten var olan filo). Yeni inşa için gerçek ömür."),

    ("c_p", "Yıllık ortalama kapasite faktörü",
     "– (0–1)", "TECHNOLOGIES",
     "NUCLEAR=0.849, PV=1(nominal), CCGT=0.85 ...",
     "PV/RES için c_p=1 set edilir; gerçek kapasite faktörü c_p_t ile tanımlanır."),

    ("fmin_perc / fmax_perc", "Teknoloji kullanım oranının alt/üst sınırı (toplam içinde)",
     "– (oran)", "TECHNOLOGIES",
     "CAR_GASOLINE: min=0.2, max=1.0 ...",
     "Politika kısıtları için kullanılır (ör. en az %20 fosil araç zorunluluğu kaldırılabilir)."),

    ("f_min / f_max", "Kurulu kapasite alt/üst sınırı",
     "GW", "TECHNOLOGIES",
     "HYDRO_DAM: min=8.08, max=8.08 (sabit); CCGT: min=0, max=10 ...",
     "Mevcut tesisler için min=max=mevcut_kapasite. Yeni inşa için f_min=0, f_max=potansiyel."),

    ("c_p_t", "Aylık kapasite faktörü (PV, RES, hidro için)",
     "– (0–1)", "TECHNOLOGIES × PERIODS",
     "PV: Ocak=0.053, Temmuz=0.164 ...",
     "EN KRİTİK parametre. NASA POWER veya Renewables.ninja'dan ülkeye özel ver."),

    # Depolama
    ("storage_eff_in / storage_eff_out", "Depolama giriş/çıkış verimliliği",
     "– (oran)", "STORAGE_TECH × LAYERS",
     "PUMPED_HYDRO: ELECTRICITY'de 1",
     "Yeni batarya için: storage_eff_in[BATTERY, ELECTRICITY]=0.95 gibi."),

    ("loss_coeff", "Katman kayıp katsayısı (iletim kayıpları)",
     "– (oran)", "LAYERS",
     "ELECTRICITY=0.07, HEAT_LOW_T_DHN=0.05",
     "Elektrik iletim kaybı ülkeye göre değişir. Türkiye ~0.12 civarı."),

    ("peak_dhn_factor", "DHN pik ısıtma/ortalama talep oranı",
     "– (oran)", "–",
     "2",
     "Soğuk iklimlerde yüksek tutulabilir (2.5–3)."),
]

for i, row in enumerate(params_data, start=3):
    bg = C_GREY if i % 2 == 0 else C_WHITE
    for j, val in enumerate(row, 1):
        data_cell(ws3, i, j, val, bg=bg)

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 45
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 40
ws3.column_dimensions["F"].width = 55
for r in range(3, 3+len(params_data)):
    ws3.row_dimensions[r].height = 50

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 – TALEP_VERISI
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("TALEP_VERISI")
ws4.sheet_view.showGridLines = False
section_title(ws4, 1, 1, "end_uses_demand_year – Sektörel Yıllık Enerji Talebi [GWh/yıl]", 8)

header_cell(ws4, 2, 1, "Açıklama", bg=C_HEADER_MED)
ws4.merge_cells("B2:H2")
c = ws4.cell(row=2, column=2,
             value="Her hücre: o nihai kullanım kategorisinin o sektördeki yıllık toplam talebini (GWh) verir. "
                   "Ulaştırma sektörü için Mpkm/Mtkm kullanılır (enerji değil, aktivite düzeyi).")
c.font      = make_font(italic=True, size=9)
c.fill      = make_fill(C_HEADER_LIGHT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.row_dimensions[2].height = 30

cols4 = ["Nihai Kullanım\n(END_USE_INPUT)", "Sektör\nAçıklaması",
         "HOUSEHOLDS\n(GWh)", "SERVICES\n(GWh)",
         "INDUSTRY\n(GWh)", "TRANSPORTATION\n(GWh/Mpkm/Mtkm)",
         "İsviçre\nBaseline Değeri", "Kaynak / Not"]
for j, h in enumerate(cols4, 1):
    header_cell(ws4, 3, j, h, bg=C_HEADER_DARK)

demand_data = [
    ("ELECTRICITY", "Doğrudan elektrik kullanımı (aydınlatma ve ısıtma hariç)\nÖrn: elektrikli cihazlar, motorlar",
     "10848.1", "15026.5", "10443.5", "0.0",
     "Toplam: 36318.1 GWh", "IEA Electricity Statistics"),
    ("LIGHTING", "Aydınlatma elektrik talebi\n(aylık dağılım: lighting_month parametresiyle)",
     "425.1", "3805.2", "1263.8", "0.0",
     "Toplam: 5494.1 GWh", "IEA; binaların aydınlatma enerji yoğunluğu"),
    ("HEAT_HIGH_T", "Yüksek sıcaklık ısı talebi (>100°C)\nSadece sanayi için geçerli",
     "0.0", "0.0", "19021.5", "0.0",
     "Sadece sanayi: 19021.5 GWh", "Proses ısısı istatistikleri"),
    ("HEAT_LOW_T_SH", "Düşük sıcaklık uzay ısıtma talebi (space heating)\nKonut+Hizmet+Sanayi",
     "29489.2", "14524.8", "4947.5", "0.0",
     "Toplam: 48961.5 GWh", "Derece-gün bazlı bina enerji modeli"),
    ("HEAT_LOW_T_HW", "Sıcak kullanma suyu talebi (hot water)\nKonut+Hizmet+Sanayi",
     "7537.8", "3256.0", "1281.8", "0.0",
     "Toplam: 12075.6 GWh", "Bina kullanım istatistikleri"),
    ("MOBILITY_PASSENGER", "Yolcu ulaşım aktivitesi\n[GWh DEĞİL → Mpkm = milyon yolcu-km]",
     "0.0", "0.0", "0.0", "146049.3 Mpkm",
     "146049.3 Mpkm", "Ulaştırma Bakanlığı, EUROSTAT"),
    ("MOBILITY_FREIGHT", "Yük taşımacılığı aktivitesi\n[GWh DEĞİL → Mtkm = milyon ton-km]",
     "0.0", "0.0", "0.0", "39966.7 Mtkm",
     "39966.7 Mtkm", "Ulaştırma Bakanlığı"),
]

for i, row in enumerate(demand_data, start=4):
    bg = C_GREEN if i % 2 == 0 else C_WHITE
    for j, val in enumerate(row, 1):
        data_cell(ws4, i, j, val, bg=bg)
    ws4.row_dimensions[i].height = 48

section_title(ws4, 12, 1, "YENİ ÜLKE İÇİN TALEP VERİSİ NASIL HAZIRLANIR?", 8, bg=C_HEADER_MED)
steps = [
    "1. IEA World Energy Balances veya ülkenin Enerji Bakanlığı istatistiklerinden sektörel enerji tüketimini al.",
    "2. Elektrik: Net elektrik tüketimini konut/hizmet/sanayi olarak ayır.",
    "3. Isıtma: Bina enerji bilançosundan uzay ısıtma (SH) ve sıcak su (HW) ayrımını yap.",
    "4. Sanayi proses ısısı: >100°C olan sanayi ısı talebini HEAT_HIGH_T'ye yaz.",
    "5. Ulaştırma: Yıllık yolcu-km ve ton-km verilerini al (GWh değil, aktivite birimi!)",
    "6. Aydınlatma payını elektrik tüketiminden ayır (tipik %10-15 konut, %20-25 hizmet).",
    "Türkiye örneği (yaklaşık, 2019): ELECTRICITY≈170000 GWh toplam, HEAT_LOW_T_SH≈180000 GWh",
]
for i, s in enumerate(steps, start=13):
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    c = ws4.cell(row=i, column=1, value=s)
    c.font      = make_font(size=9, bold=(i==13 or i==19))
    c.fill      = make_fill(C_ORANGE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[i].height = 22

for col, w in zip("ABCDEFGH", [28, 30, 12, 12, 12, 18, 20, 35]):
    ws4.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 – LAYERS_IN_OUT (özet)
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("LAYERS_IN_OUT")
ws5.sheet_view.showGridLines = False
section_title(ws5, 1, 1, "layers_in_out – Teknoloji Girdi/Çıktı Katsayıları (Verimlilik Matrisi)", 6)

ws5.merge_cells("A2:F2")
c = ws5.cell(row=2, column=1,
             value="Pozitif değer = o katmana çıktı (üretim). Negatif değer = o katmandan girdi (tüketim). "
                   "Değerin mutlak büyüklüğü, 1 GW kapasiteden kaç GWh üretildiğini/tüketildiğini gösterir.")
c.font      = make_font(italic=True, size=9)
c.fill      = make_fill(C_HEADER_LIGHT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws5.row_dimensions[2].height = 30

h5 = ["Teknoloji", "Katman (Layer)", "Değer", "Birim", "Açıklama", "Verimlilik Hesabı"]
for j, h in enumerate(h5, 1):
    header_cell(ws5, 3, j, h, bg=C_HEADER_DARK)

layers_data = [
    # Kaynaklar (pass-through)
    ("ELECTRICITY (kaynak)", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Dışarıdan alınan elektrik doğrudan sisteme girer.", "Verimlilik=1 (kayıpsız)"),
    ("COAL (kaynak)", "COAL", "+1", "GWh_kömür/GWh_in", "Kömür, COAL katmanına +1 ile girer. Teknolojiler bu katmandan çeker.", "Kaynak tanımı"),
    ("ELEC_EXPORT (ihracat)", "ELECTRICITY", "−1", "GWh/GWh", "Elektrik ihraç edildiğinde sistemden eksilir.", "İhracat = negatif tüketim"),

    # Nükleer
    ("NUCLEAR", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Çıktı: elektrik üretir", "η = 1/2.7027 = %37"),
    ("NUCLEAR", "URANIUM", "−2.7027", "GWh_uranyum/GWh_elec", "Girdi: uranyum tüketir", "2.7027 GWh uranyum → 1 GWh elektrik"),

    # CCGT
    ("CCGT", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Çıktı: elektrik", "η = 1/1.5873 = %63"),
    ("CCGT", "NG", "−1.5873", "GWh_doğalgaz/GWh_elec", "Girdi: doğal gaz", ""),

    # CCGT_CCS
    ("CCGT_CCS", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Çıktı: elektrik, CCS ile CO2 tutma", "η = 1/1.7544 = %57"),
    ("CCGT_CCS", "NG_CCS", "−1.7544", "GWh_NG_CCS/GWh_elec", "Girdi: CCS uyumlu doğal gaz", ""),

    # Kömür teknolojileri
    ("COAL_US", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Kömür buhar türbini (ultra-süperkritik)", "η = 1/2.0408 = %49"),
    ("COAL_US", "COAL", "−2.0408", "GWh_kömür/GWh_elec", "Girdi: kömür", ""),
    ("COAL_IGCC", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Entegre gazlaştırmalı kombine çevrim", "η = 1/1.8519 = %54"),
    ("COAL_IGCC", "COAL", "−1.8519", "GWh_kömür/GWh_elec", "Girdi: kömür", ""),
    ("COAL_US_CCS", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Kömür US + CCS (CO2 yakalama)", "η = 1/2.381 = %42"),
    ("COAL_US_CCS", "COAL_CCS", "−2.3810", "GWh_COAL_CCS/GWh_elec", "Girdi: CCS uyumlu kömür", ""),
    ("COAL_IGCC_CCS", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "IGCC + CCS", "η = 1/2.0833 = %48"),
    ("COAL_IGCC_CCS", "COAL_CCS", "−2.0833", "GWh_COAL_CCS/GWh_elec", "Girdi: CCS uyumlu kömür", ""),
    ("IND_BOILER_COAL", "HEAT_HIGH_T", "+1", "GWh_ısı/GWh_in", "Sanayi kömür kazanı, yüksek ısı üretir", "η = 1/1.2195 = %82"),
    ("IND_BOILER_COAL", "COAL", "−1.2195", "GWh_kömür/GWh_ısı", "Girdi: kömür", ""),

    # PV ve Rüzgar
    ("PV", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "PV güneş paneli, elektrik üretir. Girdi yok.", "Kapasite faktörü c_p_t ile verilir"),
    ("WIND", "ELECTRICITY", "+1", "GWh_elec/GWh_in", "Rüzgar türbini, elektrik üretir. Girdi yok.", "Kapasite faktörü c_p_t ile verilir"),

    # Kojenerasyon (örnek)
    ("IND_COGEN_GAS", "ELECTRICITY", "+0.9565", "GWh_elec/GWh_in", "Kojenerasyon: elektrik çıktısı", "Toplam verim ~%87"),
    ("IND_COGEN_GAS", "NG", "−2.1739", "GWh_NG/GWh_elec", "Girdi: doğal gaz", ""),
    ("IND_COGEN_GAS", "HEAT_HIGH_T", "+1", "GWh_ısı/GWh_in", "Kojenerasyon: ısı çıktısı", ""),

    # Araçlar (örnek)
    ("CAR_GASOLINE", "MOB_PRIVATE", "+1", "Mpkm/GWh_in", "Benzinli araç, özel ulaşım sağlar", ""),
    ("CAR_GASOLINE", "GASOLINE", "−0.4297", "GWh_benzin/Mpkm", "Girdi: benzin tüketimi", ""),
    ("CAR_BEV", "MOB_PRIVATE", "+1", "Mpkm/GWh_in", "Elektrikli araç", ""),
    ("CAR_BEV", "ELECTRICITY", "−0.1066", "GWh_elec/Mpkm", "Girdi: elektrik (çok daha verimli)", ""),
]

for i, row in enumerate(layers_data, start=4):
    bg = C_RED_LIGHT if "COAL" in row[0] else (C_GREEN if "PV" in row[0] or "WIND" in row[0]
         else (C_GREY if i % 2 == 0 else C_WHITE))
    for j, val in enumerate(row, 1):
        data_cell(ws5, i, j, val, bg=bg)
    ws5.row_dimensions[i].height = 30

for col, w in zip("ABCDEF", [22, 18, 8, 20, 45, 35]):
    ws5.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 – TEKNOLOJI_PARAMS
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("TEKNOLOJI_PARAMS")
ws6.sheet_view.showGridLines = False
section_title(ws6, 1, 1, "Teknoloji Parametreleri – ref_size, c_inv, c_maint, gwp_constr, lifetime, c_p, f_min, f_max", 10)

h6 = ["Teknoloji", "Kategori", "ref_size\n[GW]",
      "c_inv\n[MCHF/GW]", "c_maint\n[MCHF/GW/yıl]",
      "gwp_constr\n[tCO2/GW]", "lifetime\n[yıl]", "c_p\n[0-1]",
      "f_min\n[GW]", "f_max\n[GW]"]
for j, h in enumerate(h6, 1):
    header_cell(ws6, 2, j, h, bg=C_HEADER_DARK, size=9)

tech_params = [
    # Elektrik üretim
    ("NUCLEAR",       "Nükleer santral",           1,      5174.76, 109.92,  707.88,  60, 0.849, 0,    10),
    ("CCGT",          "Kombine çevrim (NG)",        0.5,    824.41,  21.07,   183.79,  25, 0.850, 0,    10),
    ("CCGT_CCS",      "CCGT + CO2 yakalama",        0.5,    1273.26, 30.23,   183.79,  25, 0.850, 0,    10),
    ("COAL_US",       "Kömür buhar türbini (USC)",  0.5,    2687.59, 31.72,   331.60,  35, 0.868, 0,    10),
    ("COAL_IGCC",     "Kömür IGCC",                 0.5,    3466.20, 52.27,   331.60,  35, 0.856, 0,    10),
    ("COAL_US_CCS",   "Kömür USC + CCS",            0.5,    4327.26, 67.58,   331.60,  35, 0.868, 0,    10),
    ("COAL_IGCC_CCS", "Kömür IGCC + CCS",           0.5,    6044.78, 73.86,   331.60,  35, 0.856, 0,    10),
    ("PV",            "Güneş fotovoltaik",          0.000003, 1000,  15.88,  2081.43,  25, 1,     0,    25),
    ("WIND",          "Rüzgar türbini",             0.003,  1465.62, 22.90,   622.85,  20, 1,     0,    5.3),
    ("HYDRO_DAM",     "Baraj HES (mevcut)",         8.08,   4828.39, 24.14,  1692.88,  40, 1,    8.08, 8.08),
    ("NEW_HYDRO_DAM", "Baraj HES (yeni)",           0.001,  3437.12, 2.89,   1692.88,  40, 1,     0,   0.44),
    ("GEOTHERMAL",    "Jeotermal elektrik",         0.008,  11464.13,465.04,24929.09,  30, 0.86,  0,    0.7),
    # Isıtma
    ("IND_COGEN_GAS",  "Sanayi kojenerasyon NG",   0.02,   1503.64, 98.90,  1024.32,  25, 0.85,  0,   20),
    ("IND_BOILER_GAS", "Sanayi kazan (NG)",        0.01,   62.89,   1.26,    12.32,   17, 0.95,  0,   20),
    ("IND_BOILER_COAL","Sanayi kazan (kömür)",     0.001,  123.00,  2.46,    48.18,   17, 0.90,  0,   20),
    ("DHN_BOILER_GAS", "DHN kazan (NG)",           0.01,   62.89,   1.26,    12.32,   17, 0.95,  0,   20),
    ("DEC_HP_ELEC",    "Merkezi olmayan ısı pompası",0.00001,525.45, 22.48,  164.89,  18, 0.285, 0,   20),
    ("DEC_BOILER_GAS", "Merkezi olmayan NG kazan", 0.00001,169.30,  5.08,    21.09,   17, 0.285, 0,   20),
    ("DEC_SOLAR",      "Güneş ısı kolektörü",      0.00001,767.90,  8.64,   221.22,   20, 1,     0,   20),
    # Ulaştırma
    ("CAR_GASOLINE",  "Benzinli otomobil",          0.001,  0,       0,       0,       1,  1,     0,   16.7),
    ("CAR_BEV",       "Elektrikli otomobil",        0.001,  0,       0,       0,       1,  1,     0,   16.7),
    ("CAR_FUEL_CELL", "Hidrojenli otomobil",        0.001,  0,       0,       0,       1,  1,     0,   16.7),
    # Altyapı
    ("DHN",           "Bölgesel ısıtma şebekesi",  0,      881.96,  0,       0,       60, 1,     0,   20),
    ("GRID",          "Elektrik şebekesi",          0,      61100,   0,       0,       80, 1,     0,   20),
    ("H2_ELECTROLYSIS","H2 elektroliz",             0,      328.47,  32.85,   0,       15, 0.9,   0,   20),
    # Depolama
    ("PUMPED_HYDRO",  "Pompalı depolama (HES)",     0.001,  0,       0,       0,       1,  1,     0, 2400),
    ("POWER2GAS",     "Power-to-Gas depolama",      0.001,  0.4144,  0,       0,       25, 1,     0, 100000),
]

for i, row in enumerate(tech_params, start=3):
    if "COAL" in row[0]:
        bg = C_RED_LIGHT
    elif row[0] in ("PV", "WIND", "GEOTHERMAL", "NEW_HYDRO_DAM", "DEC_SOLAR"):
        bg = C_GREEN
    elif i % 2 == 0:
        bg = C_GREY
    else:
        bg = C_WHITE
    for j, val in enumerate(row, 1):
        align = "right" if j > 2 else "left"
        data_cell(ws6, i, j, str(val), bg=bg, align=align)
    ws6.row_dimensions[i].height = 22

# Açıklama satırı
legend_row = 3 + len(tech_params) + 1
ws6.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=10)
c = ws6.cell(row=legend_row, column=1,
             value="Renk: Kırmızı=Kömür teknolojileri | Yeşil=Yenilenebilir | Beyaz/Gri=Diğer")
c.font = make_font(italic=True, size=9, color="444444")
c.fill = make_fill(C_YELLOW)
c.alignment = Alignment(horizontal="left", vertical="center")

for col, w in zip("ABCDEFGHIJ", [20, 28, 10, 12, 14, 14, 10, 8, 8, 8]):
    ws6.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 – KAYNAK_VERISI
# ═══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("KAYNAK_VERISI")
ws7.sheet_view.showGridLines = False
section_title(ws7, 1, 1, "Kaynak Verileri – avail, gwp_op, c_op (tüm dönemler sabit varsayım)", 8)

h7 = ["Kaynak", "Açıklama", "avail\n[GWh/yıl]",
      "gwp_op\n[tCO2eq/MWh]", "c_op\n[MCHF/GWh]",
      "Aylık c_op\nSabit mi?", "Türkiye için\nÖnerilen avail", "Not"]
for j, h in enumerate(h7, 1):
    header_cell(ws7, 2, j, h, bg=C_HEADER_DARK)

resources_data = [
    ("ELECTRICITY", "Şebekeden elektrik alımı (ithalat)",
     "1000000", "0.4818", "0.0901", "Evet",
     "1000000 (sınırsız)", "Grid emisyon faktörü ülkeye göre değişir"),
    ("GASOLINE", "Benzin (fosil)",
     "1000000", "0.3448", "0.0880", "Evet",
     "1000000", "Rafineri çıktısı / ithalat"),
    ("DIESEL", "Dizel yakıt",
     "1000000", "0.3148", "0.0852", "Evet",
     "1000000", ""),
    ("BIOETHANOL", "Biyoetanol (biyoyakıt)",
     "0", "0", "0.0880", "Evet",
     "0 → yerli üretim varsa güncelle", "avail=0 → mevcut değil"),
    ("BIODIESEL", "Biyodizel",
     "0", "0", "0.0852", "Evet",
     "0 → güncelle", ""),
    ("LFO", "Hafif fuel oil (kalorifer yağı)",
     "1000000", "0.3115", "0.0606", "Evet",
     "1000000", ""),
    ("LNG", "Sıvılaştırılmış doğal gaz",
     "0", "0", "0", "Evet",
     "1000000 (Türkiye LNG terminali var!)", "LNG için c_op ekle ~0.04"),
    ("NG", "Doğal gaz",
     "1000000", "0.2666", "0.0348", "Evet",
     "1000000", "Türkiye boru hattı + LNG ile alıyor"),
    ("NG_CCS", "Doğal gaz (CCS entegreli)",
     "1000000", "0.02667", "0.0348", "Evet",
     "1000000", "CCS mevcut değilse avail=0 yapılabilir"),
    ("SNG", "Sentetik doğal gaz (biyogaz)",
     "0", "0", "0.0348", "Evet",
     "0 → yerli potansiyele göre", ""),
    ("WOOD", "Odun / biyokütle",
     "12279", "0.0118", "0.0932", "Evet",
     "~50000 GWh (Türkiye ormancılık pot.)", "OGM verisinden alınabilir"),
    ("COAL", "Kömür (linyit + taş kömürü) ★",
     "1000000", "0.4014", "0.0302", "Evet",
     "1000000 (Türkiye önemli linyit kaynağı!)", "★ EN ÖNEMLİ: Türkiye'de mevcut, avail yüksek tutulabilir"),
    ("COAL_CCS", "Kömür (CCS entegreli)",
     "1000000", "0.0401", "0.0302", "Evet",
     "0 (CCS Türkiye'de henüz yok)", "Gelecek senaryoları için açık bırak"),
    ("URANIUM", "Uranyum (nükleer yakıt)",
     "1000000", "0.0039", "0.0041", "Evet",
     "1000000 (Akkuyu için)", "Akkuyu NPP devreye girince f_min güncellenmeli"),
    ("WASTE", "Atık (çöp yakma)",
     "11142", "0.1501", "0", "Evet",
     "~20000 GWh (Türkiye kentsel atık)", "TÜİK atık istatistiklerinden"),
    ("H2", "Yeşil/mavi hidrojen (dışarıdan alım)",
     "0", "0", "0", "Evet",
     "0 → elektroliz ile üretim tercih edilir", ""),
    ("ELEC_EXPORT", "Elektrik ihracatı (sanal kaynak)",
     "1000000", "0", "0", "Evet",
     "1000000", "f_max ile sınırlandır"),
]

for i, row in enumerate(resources_data, start=3):
    if "COAL" in row[0]:
        bg = C_RED_LIGHT
    elif row[0] in ("PV", "WIND", "BIOETHANOL", "BIODIESEL", "SNG", "WOOD", "WASTE", "H2"):
        bg = C_GREEN
    elif i % 2 == 0:
        bg = C_GREY
    else:
        bg = C_WHITE
    for j, val in enumerate(row, 1):
        data_cell(ws7, i, j, val, bg=bg)
    ws7.row_dimensions[i].height = 35

for col, w in zip("ABCDEFGH", [14, 28, 14, 16, 14, 12, 28, 45]):
    ws7.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8 – ZAMAN_SERISİ
# ═══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("ZAMAN_SERISİ")
ws8.sheet_view.showGridLines = False
section_title(ws8, 1, 1, "Zaman Serisi Parametreleri – period_duration, lighting_month, heating_month, c_p_t", 15)

# period_duration / lighting_month / heating_month
h8a = ["Dönem\n(Ay)", "Ay Adı", "period_duration\n[saat]",
       "lighting_month\n[oran, Σ=1]", "heating_month\n[oran, Σ=1]",
       "Açıklama", "Türkiye için\nheating_month önerisi"]
for j, h in enumerate(h8a, 1):
    header_cell(ws8, 2, j, h, bg=C_HEADER_DARK)

time_data = [
    (1,  "Ocak",    744, 0.1240, 0.1984, "En uzun geceleri olan ay. Yüksek ısıtma talebi.",  "~0.18–0.22"),
    (2,  "Şubat",   672, 0.1020, 0.1654, "Kısa ay (28/29 gün = 672/696 saat).",              "~0.15–0.18"),
    (3,  "Mart",    744, 0.0918, 0.1421, "İlkbahar başı, ısıtma azalmaya başlar.",            "~0.12–0.15"),
    (4,  "Nisan",   720, 0.0708, 0.0319, "Ilıman hava, düşük ısıtma.",                        "~0.05–0.08"),
    (5,  "Mayıs",   744, 0.0588, 0,      "Isıtma yok, serin geçen günler.",                  "0"),
    (6,  "Haziran", 720, 0.0438, 0,      "Yaz başı, soğutma talebi başlar (modelde yok).",   "0"),
    (7,  "Temmuz",  744, 0.0408, 0,      "En kısa geceleri. Maksimum PV üretim.",            "0"),
    (8,  "Ağustos", 744, 0.0398, 0,      "Yüksek sıcaklık, düşük ısıtma.",                  "0"),
    (9,  "Eylül",   720, 0.0678, 0.0147, "Sonbahar başı.",                                   "~0.01–0.03"),
    (10, "Ekim",    744, 0.1048, 0.0898, "Gündüzler kısalır, ısıtma başlar.",                "~0.08–0.12"),
    (11, "Kasım",   720, 0.1238, 0.1383, "Kış yaklaşır.",                                    "~0.14–0.17"),
    (12, "Aralık",  744, 0.1318, 0.2194, "En yüksek ısıtma talebi.",                         "~0.18–0.22"),
]

for i, row in enumerate(time_data, start=3):
    bg = C_GREEN if row[4] == 0 else (C_GREY if i % 2 == 0 else C_WHITE)
    for j, val in enumerate(row, 1):
        data_cell(ws8, i, j, str(val), bg=bg)
    ws8.row_dimensions[i].height = 24

# c_p_t tablosu
section_title(ws8, 16, 1, "c_p_t – Aylık Kapasite Faktörleri (İsviçre Verisi)", 15, bg=C_HEADER_MED)
techs_cpt = ["PV", "WIND", "HYDRO_DAM", "NEW_HYDRO_DAM", "HYDRO_RIVER", "NEW_HYDRO_RIVER", "DEC_SOLAR"]
month_names = ["Oca","Şub","Mar","Nis","May","Haz","Tem","Ağu","Eyl","Eki","Kas","Ara"]
cpt_header = ["Teknoloji"] + month_names + ["Açıklama"]
for j, h in enumerate(cpt_header, 1):
    header_cell(ws8, 17, j, h, bg=C_HEADER_DARK, size=9)

cpt_vals = [
    ("PV",           [0.053,0.094,0.120,0.153,0.156,0.157,0.164,0.156,0.128,0.088,0.052,0.036],
     "Güneş radyasyonu. Türkiye için Temmuz daha yüksek (0.20+) olabilir."),
    ("WIND",         [0.325,0.279,0.282,0.182,0.176,0.140,0.149,0.135,0.162,0.279,0.313,0.339],
     "Rüzgar kf. Türkiye için Ege/Marmara kıyısı: kış daha yüksek."),
    ("HYDRO_DAM",    [0.251,0.253,0.188,0.160,0.212,0.279,0.268,0.259,0.288,0.212,0.217,0.226],
     "Baraj HES: depolama sayesinde daha düzgün profil."),
    ("NEW_HYDRO_DAM",[0.251,0.253,0.188,0.160,0.212,0.279,0.268,0.259,0.288,0.212,0.217,0.226],
     "Aynı profil, yeni baraj kapasitesi için."),
    ("HYDRO_RIVER",  [0.302,0.265,0.308,0.436,0.622,0.766,0.770,0.713,0.548,0.376,0.348,0.342],
     "Nehir HES: kar erimesiyle Nisan-Temmuz yüksek."),
    ("NEW_HYDRO_RIVER",[0.302,0.265,0.308,0.436,0.622,0.766,0.770,0.713,0.548,0.376,0.348,0.342],
     "Aynı profil, yeni nehir HES için."),
    ("DEC_SOLAR",    [0.065,0.091,0.120,0.123,0.141,0.152,0.161,0.149,0.135,0.101,0.068,0.055],
     "Termal güneş kolektörü (ısıtma), PV'den biraz farklı profil."),
]

for i, (tech, vals, note) in enumerate(cpt_vals, start=18):
    bg = C_GREEN if i % 2 == 0 else C_WHITE
    ws8.cell(row=i, column=1, value=tech).font = make_font(bold=True)
    ws8.cell(row=i, column=1).fill = make_fill(bg)
    ws8.cell(row=i, column=1).border = make_border()
    for j, v in enumerate(vals, 2):
        data_cell(ws8, i, j, v, bg=bg, align="center")
    data_cell(ws8, i, 14, note, bg=bg)
    ws8.row_dimensions[i].height = 22

note_row = 18 + len(cpt_vals) + 1
ws8.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=15)
c = ws8.cell(row=note_row, column=1,
             value="⚠ ÖNEMLI: Türkiye için bu değerleri NASA POWER (power.larc.nasa.gov) veya "
                   "Renewables.ninja (renewables.ninja) sitesinden ülkeye özel koordinatlarla güncelle!")
c.font      = make_font(bold=True, color="CC0000", size=10)
c.fill      = make_fill(C_YELLOW)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws8.row_dimensions[note_row].height = 28

for col, w in zip("ABCDEFGHIJKLMNO", [16,7,7,7,7,7,7,7,7,7,7,7,7,7,40]):
    ws8.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9 – DEPOLAMA
# ═══════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("DEPOLAMA")
ws9.sheet_view.showGridLines = False
section_title(ws9, 1, 1, "Depolama Parametreleri – storage_eff_in/out, loss_coeff, peak_dhn_factor", 6)

h9 = ["Parametre", "Teknoloji / Katman", "Değer", "Birim", "Açıklama", "Yeni Batarya Eklemek İçin"]
for j, h in enumerate(h9, 1):
    header_cell(ws9, 2, j, h, bg=C_HEADER_DARK)

storage_rows = [
    ("storage_eff_in", "PUMPED_HYDRO → ELECTRICITY", "1.0",
     "–", "Pompalı depolamaya elektrik girişinde kayıp yok (idealleştirilmiş).",
     "BATTERY → ELECTRICITY: 0.95 kullan"),
    ("storage_eff_out", "PUMPED_HYDRO → ELECTRICITY", "1.0",
     "–", "Çıkışta da kayıp yok (gerçekte round-trip ~%75-85).",
     "BATTERY → ELECTRICITY: 0.95 kullan"),
    ("storage_eff_in", "POWER2GAS → LNG", "1.0",
     "–", "Power-to-Gas: elektrik → LNG dönüşüm girişi",
     "BATTERY → ELECTRICITY: 0.95 kullan"),
    ("storage_eff_out", "POWER2GAS → LNG", "1.0",
     "–", "Power-to-Gas: LNG → elektrik çıkışı",
     ""),
    ("loss_coeff", "ELECTRICITY", "0.07",
     "–", "Elektrik iletim kayıp oranı (%7).",
     "Türkiye 2023 iletim kaybı ~%12 → 0.12 yap"),
    ("loss_coeff", "HEAT_LOW_T_DHN", "0.05",
     "–", "Bölgesel ısıtma ağı dağıtım kayıpları (%5).",
     "Eski şebekede %10-15 olabilir → 0.10 yap"),
    ("peak_dhn_factor", "–", "2",
     "–", "DHN pik talebi / ortalama talep oranı. Kapasite boyutlandırması için.",
     "Soğuk iklimde 2.5-3 yapılabilir"),
]

for i, row in enumerate(storage_rows, start=3):
    bg = C_GREY if i % 2 == 0 else C_WHITE
    for j, val in enumerate(row, 1):
        data_cell(ws9, i, j, val, bg=bg)
    ws9.row_dimensions[i].height = 35

# Yeni batarya ekleme kutusu
box_row = 3 + len(storage_rows) + 2
section_title(ws9, box_row, 1, "BATARYA DEPOLAMA EKLEME (BATTERY) – Adımlar", 6, bg=C_HEADER_MED)
battery_steps = [
    "1. SETS'e ekle: set STORAGE_TECH := PUMPED_HYDRO POWER2GAS BATTERY;",
    "2. storage_eff_in'e satır ekle:   BATTERY 0.95 0 0 0 0 0 ...  (ELECTRICITY sütunu 0.95)",
    "3. storage_eff_out'a satır ekle:  BATTERY 0.95 0 0 0 0 0 ...",
    "4. tech_params'a satır ekle:      BATTERY  0  500  15  0  15  1  0  0  0  10000",
    "   (ref_size=0, c_inv=500 MCHF/GWh, c_maint=15, gwp=0, lifetime=15, c_p=1, f_min=0, f_max=10000)",
    "5. layers_in_out'a satır ekle:    BATTERY  -1  0  0  ...  0  (ELECTRICITY=-1 girdi)",
    "   VE:                             BATTERY   1  0  0  ...  0  (ELECTRICITY=+1 çıktı olarak ayrı satır)",
    "   NOT: Modelin matematiksel yapısı storage için ayrı değişkenler kullanır (storage_level)",
]
for i, s in enumerate(battery_steps, start=box_row+1):
    ws9.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    c = ws9.cell(row=i, column=1, value=s)
    c.font      = make_font(size=9, bold=s.startswith("   NOT"))
    c.fill      = make_fill(C_ORANGE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws9.row_dimensions[i].height = 22

for col, w in zip("ABCDEF", [22, 25, 8, 8, 45, 40]):
    ws9.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10 – YENI_ULKE_REHBERI
# ═══════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("YENI_ULKE_REHBERI")
ws10.sheet_view.showGridLines = False

ws10.merge_cells("A1:C1")
c = ws10.cell(row=1, column=1, value="YENİ BİR ÜLKE MODELLEMEK İÇİN ADIM ADIM KILAVUZ")
c.font      = make_font(bold=True, color="FFFFFF", size=14)
c.fill      = make_fill(C_HEADER_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws10.row_dimensions[1].height = 36

steps_country = [
    ("ADIM 1", "Proje Klasörü Hazırla",
     "scenarios/<ulke_adi>/ klasörü oluştur.\n"
     "Mevcut scenarios/baseline/ içeriğini kopyala başlangıç noktası olarak kullan:\n"
     "  cp -r scenarios/baseline scenarios/turkey_2030\n"
     "Değiştirilecek dosya: ses_main.dat"),

    ("ADIM 2", "SETS – Genellikle değiştirme gerekmez",
     "PERIODS, SECTORS, END_USES_INPUT, RESOURCES, TECHNOLOGIES setleri aynı kalabilir.\n"
     "Eğer yeni teknoloji veya kaynak ekliyorsan:\n"
     "  → RESOURCES setine yeni kaynağı ekle (örn. LIGNITE)\n"
     "  → TECHNOLOGIES_OF_END_USES_TYPE[...] setine yeni teknolojiyi ekle"),

    ("ADIM 3", "i_rate – İskonto Oranı",
     "Ülkenin risk-free faiz oranını kullan.\n"
     "Türkiye örneği: param i_rate := 0.08;\n"
     "Kaynak: Merkez Bankası uzun vadeli devlet tahvil faizi"),

    ("ADIM 4", "end_uses_demand_year – EN KRİTİK",
     "Her satır: bir talep kategorisi. Her sütun: bir sektör.\n\n"
     "Türkiye 2019 yaklaşık değerleri (GWh):\n"
     "ELECTRICITY:       35000  50000  70000  0\n"
     "LIGHTING:          3000   12000  5000   0\n"
     "HEAT_HIGH_T:       0      0      80000  0\n"
     "HEAT_LOW_T_SH:     60000  25000  15000  0\n"
     "HEAT_LOW_T_HW:     12000  5000   3000   0\n"
     "MOBILITY_PASSENGER:0      0      0      300000 Mpkm\n"
     "MOBILITY_FREIGHT:  0      0      0      250000 Mtkm\n\n"
     "Kaynak: IEA Energy Balances, TEIAŞ, Ulaştırma Bakanlığı"),

    ("ADIM 5", "Ulaştırma Payları",
     "Türkiye için önerilen değerler:\n"
     "param share_mobility_public_min := 0.15;  # karayolu baskın\n"
     "param share_mobility_public_max := 0.35;\n"
     "param share_freight_train_min   := 0.05;  # demir yolu payı düşük\n"
     "param share_freight_train_max   := 0.20;\n"
     "param share_heat_dhn_min        := 0.00;\n"
     "param share_heat_dhn_max        := 0.10;"),

    ("ADIM 6", "period_duration & aylık dağılımlar",
     "Aylık çözünürlük için dönem süreleri değişmez (yılın ayları).\n\n"
     "heating_month → Türkiye için ısıtma derece-günleri:\n"
     "  İç Anadolu: Ocak %22, Şubat %18, ...\n"
     "  Ege kıyısı: Ocak %12, Şubat %10, ...\n\n"
     "lighting_month → Güneş ışığı saatlerine ters orantılı\n"
     "  Temmuz en düşük (~0.04), Aralık en yüksek (~0.13)\n\n"
     "ÖNEMLİ: Tüm aylık dağılım değerleri toplamı = 1.00 olmalı!"),

    ("ADIM 7", "layers_in_out – Verimlilik Matrisi",
     "Teknoloji verimlilikleri uluslararası standart değerler kullandığından\n"
     "genellikle değiştirme gerekmez.\n\n"
     "Kömür için mevcut teknolojiler (set RESOURCES'da COAL zaten var!):\n"
     "  COAL_US: η=%49, COAL_IGCC: η=%54\n"
     "  IND_BOILER_COAL: η=%82\n\n"
     "Yeni teknoloji ekliyorsan tek bir satır ekle:\n"
     "  YENI_TECH  +1  0  0  -verimlilik_tersi  0  ..."),

    ("ADIM 8", "avail – Kaynak Kullanılabilirliği",
     "Türkiye için önerilen değişiklikler:\n"
     "WOOD   := 50000;   # Orman potansiyeli (OGM verisi)\n"
     "WASTE  := 20000;   # Kentsel atık potansiyeli (TÜİK)\n"
     "COAL   := 1000000; # Linyit bolluğu, sınırsız bırakılabilir\n"
     "LNG    := 1000000; # Türkiye LNG terminallerine sahip\n"
     "NG     := 1000000; # Boru hatları + LNG ile alım\n"
     "BIOETHANOL := 5000; # Şeker pancarı bazlı biyoetanol\n\n"
     "Kaynak: Enerji Atlası, MTA, EPDK raporları"),

    ("ADIM 9", "gwp_op – Emisyon Faktörleri",
     "IPCC AR6 default değerleri kullanılabilir (değişiklik gerekmez).\n"
     "Türkiye grid emisyon faktörü ~0.45 tCO2/MWh (2023 verisi)\n"
     "  → ELECTRICITY gwp_op := 0.45;\n\n"
     "Kaynak: IEA Electricity Information, IPCC 2006 Guidelines"),

    ("ADIM 10", "c_op – Kaynak Maliyetleri",
     "Türkiye enerji fiyatları (yaklaşık 2023, USD/MWh → MCHF/GWh × 0.001 × kur):\n"
     "NG     := 0.055;  # Doğal gaz ~55 USD/MWh\n"
     "COAL   := 0.020;  # Linyit ucuz ~20 USD/MWh (yerli)\n"
     "DIESEL := 0.090;\n"
     "GASOLINE := 0.095;\n"
     "LNG    := 0.060;  # LNG ithalat fiyatı\n\n"
     "Not: c_op 12 dönem için aynı değer olabilir (sabit fiyat varsayımı).\n"
     "Mevsimsel fiyat farkı varsa her dönem için farklı değer girilebilir."),

    ("ADIM 11", "f_min/f_max – Mevcut ve Planlanan Kapasiteler",
     "Mevcut kurulu kapasite → f_min = f_max = mevcut_kapasite\n"
     "Yeni yatırım potansiyeli → f_min = 0, f_max = teknik_potansiyel\n\n"
     "Türkiye 2023 mevcut kapasiteler (GW):\n"
     "HYDRO_DAM:   f_min=30, f_max=30   (mevcut baraj HES)\n"
     "HYDRO_RIVER: f_min=5,  f_max=5    (nehir HES)\n"
     "PV:          f_min=0,  f_max=100  (büyük güneş potansiyeli)\n"
     "WIND:        f_min=0,  f_max=50   (karasal + deniz üstü)\n"
     "COAL_US:     f_min=10, f_max=10   (mevcut kömür santralleri)\n"
     "NUCLEAR:     f_min=0,  f_max=5    (Akkuyu 4.8 GW)\n\n"
     "Kaynak: TEİAŞ Kapasite Raporu, EPDK Lisans Dairesi"),

    ("ADIM 12", "c_p_t – Aylık Kapasite Faktörleri (ZORUNLU GÜNCELLEMEdir)",
     "Bu değerleri NASA POWER API'den veya Renewables.ninja'dan al:\n\n"
     "NASA POWER için Türkiye örnek koordinat: Lat=39, Lon=35 (İç Anadolu)\n"
     "  https://power.larc.nasa.gov/data-access-viewer/\n\n"
     "Renewables.ninja için:\n"
     "  https://www.renewables.ninja/ → Tools → API\n\n"
     "İndirilen veriden aylık ortalama hesapla:\n"
     "  PV kf: Temmuz ~0.22, Ocak ~0.10 (İzmir bölgesi)\n"
     "  Wind kf: Ocak ~0.40, Temmuz ~0.20 (Ege kıyısı)"),

    ("ADIM 13", "Doğrulama ve Model Çalıştırma",
     "1. Toplam talebi kontrol et: sum(end_uses_demand_year) makul mü?\n"
     "2. share parametrelerinin toplamı mantıklı mı?\n"
     "3. heating_month + lighting_month her biri toplamda 1.00 mi?\n"
     "4. c_p_t değerleri 0-1 arasında mı?\n"
     "5. f_min ≤ f_max her zaman sağlanmalı!\n\n"
     "Çalıştırma:\n"
     "  glpsol --math ses_model.mod --data scenarios/turkey_2030/ses_main.dat\n"
     "veya\n"
     "  ampl: data scenarios/turkey_2030/ses_main.dat; solve;"),
]

for i, (step, title, detail) in enumerate(steps_country, start=2):
    ws10.row_dimensions[i].height = 120
    c_step = ws10.cell(row=i, column=1, value=f"{step}\n{title}")
    c_step.font      = make_font(bold=True, color="FFFFFF", size=10)
    c_step.fill      = make_fill(C_HEADER_MED if i%2==0 else C_HEADER_DARK)
    c_step.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_step.border    = make_border(thick=True)

    ws10.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    c_det = ws10.cell(row=i, column=2, value=detail)
    c_det.font      = make_font(size=9)
    c_det.fill      = make_fill(C_GREY if i%2==0 else C_WHITE)
    c_det.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_det.border    = make_border()

ws10.column_dimensions["A"].width = 22
ws10.column_dimensions["B"].width = 70
ws10.column_dimensions["C"].width = 5

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11 – KOMUR_EKLEME
# ═══════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("KOMUR_EKLEME")
ws11.sheet_view.showGridLines = False

ws11.merge_cells("A1:D1")
c = ws11.cell(row=1, column=1,
              value="KÖMÜR ENERJİ KAYNAĞI / TEKNOLOJİSİ EKLEME REHBERİ")
c.font      = make_font(bold=True, color="FFFFFF", size=14)
c.fill      = make_fill(C_HEADER_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws11.row_dimensions[1].height = 36

info_row = 2
ws11.merge_cells("A2:D2")
c = ws11.cell(row=2, column=1,
              value="ÖNEMLİ NOT: COAL kaynağı ses_main.dat'ta ZATen MEVCUT! "
                    "COAL_US, COAL_IGCC, IND_BOILER_COAL teknolojileri de mevcut. "
                    "Sadece avail ve f_max değerlerini ayarlamak genellikle yeterlidir.")
c.font      = make_font(bold=True, color="CC0000", size=10)
c.fill      = make_fill(C_YELLOW)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws11.row_dimensions[2].height = 40

coal_sections = [
    ("SENARYO A", "Mevcut Kömür Teknolojisini Aktive Etmek (En Basit Yol)",
     [
         ("Dosya", "scenarios/<ulke>/ses_main.dat"),
         ("Değişiklik 1 – avail", "COAL kaynağının avail değerini kontrol et ve artır:\n"
          "  param: avail :=\n  ...\n  COAL  1000000   # zaten sınırsız"),
         ("Değişiklik 2 – c_op", "Kömür işletme maliyetini ülkeye göre ayarla:\n"
          "  param c_op: 1 2 3 4 5 6 7 8 9 10 11 12 :=\n"
          "  COAL  0.020 0.020 0.020 0.020 0.020 0.020 0.020 0.020 0.020 0.020 0.020 0.020"),
         ("Değişiklik 3 – f_max", "İstenen kömür santrali kapasitesine izin ver:\n"
          "  COAL_US   ... f_min=0  f_max=20  (max 20 GW kömür kapasitesine izin ver)\n"
          "  IND_BOILER_COAL  f_min=0 f_max=50"),
         ("Sonuç", "Model artık kömürü ekonomikse kullanacak. Sonuçta COAL_US, "
          "COAL_IGCC veya IND_BOILER_COAL teknolojileri seçilebilir."),
     ]),
    ("SENARYO B", "Yeni Kömür Teknolojisi Eklemek (Örn: Yerli Linyit Santrali LIG_PLANT)",
     [
         ("Adım 1 – RESOURCES", "COAL zaten var, ama Linyit için ayrı kaynak eklemek istersen:\n"
          "  set RESOURCES := ... COAL LIGNITE ...;\n"
          "  Sonra avail, gwp_op, c_op için LIGNITE satırı ekle."),
         ("Adım 2 – TECHNOLOGIES_OF_END_USES_TYPE", "Yeni teknolojiyi elektrik setine ekle:\n"
          "  set TECHNOLOGIES_OF_END_USES_TYPE['ELECTRICITY'] :=\n"
          "    NUCLEAR CCGT ... COAL_US LIG_PLANT;  # sona ekle"),
         ("Adım 3 – layers_in_out", "Teknolojinin girdi/çıktı matrisine satır ekle:\n"
          "  LIG_PLANT  1  0  0  0  0  0  0  0  -2.2  0  0  0  0  0  0  0  0  0  0  0\n"
          "  ↑ Açıklama: 1 GWh elektrik çıktısı, -2.2 GWh linyit girişi (η≈%45)"),
         ("Adım 4 – ref_size, c_inv, ...", "Tech parametreleri tablosuna satır ekle:\n"
          "  LIG_PLANT  0.5  2500  35  280  35  0.85  0  0.5  0  15\n"
          "  (ref_size=0.5GW, c_inv=2500 MCHF/GW, ömür=35yıl, c_p=0.85)"),
         ("Adım 5 – gwp_op for LIGNITE", "Linyit için emisyon faktörü ekle:\n"
          "  param: gwp_op :=\n  ...\n  LIGNITE  0.50   # ~500 kgCO2/MWh (linyit için yüksek)"),
         ("Adım 6 – Doğrulama", "Model çalıştır ve COAL_US vs LIG_PLANT rekabetini gözlemle.\n"
          "Linyit ucuzsa ama karbon maliyeti yüksekse seçilmeyebilir."),
     ]),
    ("SENARYO C", "Kömüre Aylık Farklı Maliyet Atamak",
     [
         ("Açıklama", "Kış aylarında kömür fiyatı yükseliyorsa c_op periyot bazında farklı ayarlanabilir:"),
         ("Örnek", "param c_op: 1    2    3    4    5    6    7    8    9    10   11   12  :=\n"
          "  COAL  0.025 0.025 0.022 0.020 0.018 0.018 0.018 0.018 0.020 0.022 0.025 0.027"),
         ("Etki", "Model kışın kömürü daha pahalı olarak değerlendirir, alternatif kaynakları tercih edebilir."),
     ]),
    ("SENARYO D", "Kömür Santrallerini Emekli Etmek / Kapatmak",
     [
         ("Açıklama", "2030'dan sonra kömür santrallerini devre dışı bırakmak için f_max=0 yap:"),
         ("Örnek (2030 senaryosu)", "COAL_US    ...  f_min=0  f_max=0   # kapatıldı\n"
          "COAL_IGCC  ...  f_min=0  f_max=0   # kapatıldı\n"
          "IND_BOILER_COAL  f_min=0  f_max=0  # kapatıldı"),
         ("Veya", "Sadece f_max'ı mevcut kapasitenin altına çek:\n"
          "COAL_US  ...  f_min=5  f_max=5   # 5 GW sabit, azaltılamaz, artırılamaz"),
     ]),
]

current_row = 3
for section_name, section_title_text, items in coal_sections:
    section_title(ws11, current_row, 1, f"{section_name}: {section_title_text}", 4)
    ws11.row_dimensions[current_row].height = 28
    current_row += 1

    for key, val in items:
        ws11.row_dimensions[current_row].height = 60
        c_key = ws11.cell(row=current_row, column=1, value=key)
        c_key.font      = make_font(bold=True, size=9)
        c_key.fill      = make_fill(C_HEADER_LIGHT)
        c_key.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        c_key.border    = make_border()

        ws11.merge_cells(start_row=current_row, start_column=2,
                         end_row=current_row, end_column=4)
        c_val = ws11.cell(row=current_row, column=2, value=val)
        c_val.font      = make_font(size=9)
        c_val.fill      = make_fill(C_WHITE)
        c_val.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c_val.border    = make_border()
        current_row += 1

    current_row += 1  # boşluk

for col, w in zip("ABCD", [20, 50, 5, 5]):
    ws11.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 12 – AYLIK_COZUNURLUK
# ═══════════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("AYLIK_COZUNURLUK")
ws12.sheet_view.showGridLines = False

ws12.merge_cells("A1:C1")
c = ws12.cell(row=1, column=1,
              value="AYLIK ÇÖZÜNÜRLÜK (12 DÖNEM) – AYARLAR VE AÇIKLAMALAR")
c.font      = make_font(bold=True, color="FFFFFF", size=14)
c.fill      = make_fill(C_HEADER_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws12.row_dimensions[1].height = 36

monthly_sections = [
    ("MEVCUT DURUM", "ses_main.dat ZATen 12 aylık çözünürlükte çalışıyor!",
     "set PERIODS := 1 2 3 4 5 6 7 8 9 10 11 12;\n\n"
     "Her dönem = 1 ay. Model bu 12 dönemi optimize ederek:\n"
     "  • Her ay için kaynak kullanımını\n"
     "  • Her ay için teknoloji çalışmasını\n"
     "  • Depolama şarj/deşarj dengesini\n"
     "  belirler."),

    ("period_duration PARAMETRESI", "Her dönemin kaç saat olduğunu söyler",
     "param: period_duration :=\n"
     "  1   744   # Ocak    (31 gün × 24 = 744 saat)\n"
     "  2   672   # Şubat   (28 gün × 24 = 672 saat)\n"
     "  3   744   # Mart\n"
     "  4   720   # Nisan   (30 gün × 24 = 720 saat)\n"
     "  5   744   # Mayıs\n"
     "  6   720   # Haziran\n"
     "  7   744   # Temmuz\n"
     "  8   744   # Ağustos\n"
     "  9   720   # Eylül\n"
     " 10   744   # Ekim\n"
     " 11   720   # Kasım\n"
     " 12   744   # Aralık\n"
     ";\n\n"
     "Artık yıl için Şubat = 696 saat yapılabilir."),

    ("lighting_month PARAMETRESI", "Aydınlatma talebinin aylara dağıtımı",
     "Değerlerin toplamı = 1.00 olmalı!\n\n"
     "Türkiye için önerilen:\n"
     "param: lighting_month :=\n"
     "  1  0.1180   # Ocak    - uzun gece\n"
     "  2  0.1000   # Şubat\n"
     "  3  0.0920   # Mart\n"
     "  4  0.0720   # Nisan\n"
     "  5  0.0600   # Mayıs\n"
     "  6  0.0450   # Haziran - kısa gece\n"
     "  7  0.0400   # Temmuz  - en kısa\n"
     "  8  0.0400   # Ağustos\n"
     "  9  0.0680   # Eylül\n"
     " 10  0.1050   # Ekim\n"
     " 11  0.1200   # Kasım\n"
     " 12  0.1400   # Aralık  - en uzun gece\n"
     ";"),

    ("heating_month PARAMETRESI", "Isıtma talebinin aylara dağıtımı",
     "Değerlerin toplamı = 1.00 olmalı (ısıtma gereken aylar için)!\n\n"
     "Türkiye iç kesimler için (HDD bazlı):\n"
     "param: heating_month :=\n"
     "  1  0.2100   # Ocak    - en soğuk\n"
     "  2  0.1700   # Şubat\n"
     "  3  0.1400   # Mart\n"
     "  4  0.0500   # Nisan   - bahar başı\n"
     "  5  0.0000   # Mayıs   - ısıtma yok\n"
     "  6  0.0000   # Haziran\n"
     "  7  0.0000   # Temmuz\n"
     "  8  0.0000   # Ağustos\n"
     "  9  0.0200   # Eylül   - erken sonbahar\n"
     " 10  0.1000   # Ekim\n"
     " 11  0.1500   # Kasım\n"
     " 12  0.1600   # Aralık\n"
     ";\n\n"
     "Kaynak: Meteoroloji Genel Müdürlüğü HDD verileri"),

    ("c_p_t PARAMETRESI", "Aylık kapasite faktörleri – sadece değişken kaynaklara uygulanır",
     "Sabit kaynaklara (NG, COAL, NUCLEAR) c_p_t GİREMEZSİN!\n"
     "Sadece şu teknolojiler için c_p_t tanımlanır:\n"
     "  PV, WIND, HYDRO_DAM, HYDRO_RIVER, DEC_SOLAR\n\n"
     "Örnek format:\n"
     "param c_p_t: 1    2    3    4    5    6    7    8    9    10   11   12  :=\n"
     "PV          0.08 0.10 0.13 0.17 0.20 0.22 0.23 0.21 0.18 0.13 0.09 0.07\n"
     "WIND        0.38 0.35 0.30 0.22 0.18 0.15 0.14 0.14 0.18 0.26 0.34 0.38\n"
     "HYDRO_DAM   0.25 0.25 0.20 0.18 0.22 0.30 0.28 0.26 0.29 0.22 0.22 0.23\n"
     ";"),

    ("c_op PARAMETRESI (aylık farklı fiyat)", "Kaynakların aylık farklı fiyatlarını girme",
     "Eğer fiyatlar mevsimsel ise her dönem farklı değer girilebilir:\n\n"
     "param c_op: 1    2    3    4    5    6    7    8    9    10   11   12 :=\n"
     "NG   0.040 0.040 0.038 0.035 0.032 0.030 0.030 0.030 0.033 0.037 0.040 0.042\n"
     ";  # Kışın doğal gaz daha pahalı\n\n"
     "Sabit fiyat varsayımı (daha basit):\n"
     "NG   0.035 0.035 0.035 0.035 0.035 0.035 0.035 0.035 0.035 0.035 0.035 0.035"),

    ("YILLIK → AYLIK ÇÖZÜNÜRLÜĞE DÖNÜŞÜM", "Eğer başka bir modelden 12 periyotlu yapıya geçiyorsan",
     "1. set PERIODS'ı 1→12 yap (zaten öyle)\n"
     "2. period_duration'a her ay için saat sayısını gir\n"
     "3. lighting_month ve heating_month toplamı 1 olacak şekilde dağıt\n"
     "4. c_p_t tablosunu güncelle (12 sütun)\n"
     "5. c_op tablosunu güncelle (12 dönem × kaynak sayısı)\n\n"
     "Model aynı; sadece veri daha ayrıntılı giriliyor.\n"
     "12 dönem → daha uzun çözüm süresi ama daha gerçekçi mevsimsel sonuçlar."),

    ("SAATLIK ÇÖZÜNÜRLÜĞE GEÇİŞ (İleri Seviye)", "8760 periyot kullanmak için",
     "1. set PERIODS := 1 2 ... 8760;\n"
     "2. period_duration: her saat için 1 gir\n"
     "3. lighting_month → saat bazlı aydınlatma profili\n"
     "4. heating_month  → saat bazlı ısıtma profili\n"
     "5. c_p_t          → saat bazlı kapasite faktörleri (PVGIS'ten)\n\n"
     "UYARI: 8760 periyot → çözüm saatler alabilir!\n"
     "Cluster yöntemi (12 tipik gün seçimi) tercih edilebilir.\n"
     "Referans: EnergyScope TD (Temporal Decomposition) yöntemi"),
]

curr_r = 2
for title_ms, subtitle_ms, content_ms in monthly_sections:
    ws12.row_dimensions[curr_r].height = 28
    section_title(ws12, curr_r, 1, f"{title_ms}: {subtitle_ms}", 3)
    curr_r += 1

    ws12.row_dimensions[curr_r].height = 130
    ws12.merge_cells(start_row=curr_r, start_column=1,
                     end_row=curr_r, end_column=3)
    c = ws12.cell(row=curr_r, column=1, value=content_ms)
    c.font      = make_font(size=9)
    c.fill      = make_fill(C_WHITE if curr_r % 4 > 1 else C_GREY)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border    = make_border()
    curr_r += 2

for col, w in zip("ABC", [28, 5, 5]):
    ws12.column_dimensions[col].width = w
ws12.column_dimensions["A"].width = 95

# ─── Son: kaydet ──────────────────────────────────────────────────────────
out_path = "/Users/eminkartci/Desktop/codenv/ECCM-Project-Group10/SES_Model_Rehberi.xlsx"
wb.save(out_path)
print(f"Kaydedildi: {out_path}")
